import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import random
import io


    
st.set_page_config(page_title="Amazon ASIN Lookup", page_icon="🛒", layout="centered")

st.markdown("""
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# Password gate


st.title("🛒 Amazon ASIN Lookup")
st.caption("Fetch product titles from Amazon.in — single ASIN or bulk CSV upload")

# ── helpers ──────────────────────────────────────────────────────────────────

def get_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service

    options = Options()
    options.binary_location = "/usr/bin/chromium"
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service("/usr/bin/chromedriver")
    return webdriver.Chrome(service=service, options=options)


def fetch_title(driver, asin: str) -> str:
    from selenium.webdriver.common.by import By
    url = f"https://www.amazon.in/dp/{asin}"
    try:
        driver.get(url)
        time.sleep(random.uniform(2, 4))
        title = driver.find_element(By.ID, "productTitle").text.strip()
        return title if title else "Title not found"
    except Exception:
        return "Title not found"


def build_excel(results: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amazon Products"

    header_fill   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
    header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    url_font      = Font(color="0563C1", underline="single", name="Arial", size=10)
    normal_font   = Font(name="Arial", size=10)
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col, h in enumerate(["ASIN", "Amazon URL", "Product Title"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    for r, row in enumerate(results, 2):
        ws.cell(row=r, column=1, value=row["ASIN"]).font = normal_font
        ws.cell(row=r, column=1).alignment = center

        uc = ws.cell(row=r, column=2, value=row["Amazon URL"])
        uc.font = url_font
        uc.alignment = left
        uc.hyperlink = row["Amazon URL"]

        tc = ws.cell(row=r, column=3, value=row["Product Title"])
        tc.font = normal_font
        tc.alignment = left

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI tabs ───────────────────────────────────────────────────────────────────

tab1, tab2 = st.tabs(["🔍 Single ASIN", "📁 Bulk CSV Upload"])

# ─── Tab 1: Single ASIN ───────────────────────────────────────────────────────
with tab1:
    st.subheader("Look up one product")
    asin_input = st.text_input("Enter ASIN", placeholder="e.g. B0CHX3QBCH", max_chars=12)

    if st.button("Fetch Title", type="primary", key="single_btn"):
        asin = asin_input.strip().upper()
        if not asin:
            st.warning("Please enter an ASIN.")
        else:
            with st.spinner("Opening Amazon page…"):
                try:
                    driver = get_driver()
                    title = fetch_title(driver, asin)
                    driver.quit()

                    url = f"https://www.amazon.in/dp/{asin}"
                    results = [{"ASIN": asin, "Amazon URL": url, "Product Title": title}]

                    if title == "Title not found":
                        st.error(f"❌ Could not find title for **{asin}**. The product may be unavailable or the page was blocked.")
                    else:
                        st.success(f"✅ **{title}**")
                        st.markdown(f"[View on Amazon.in]({url})")

                    excel_bytes = build_excel(results)
                    st.download_button(
                        label="⬇️ Download Excel",
                        data=excel_bytes,
                        file_name=f"{asin}_product.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"Error: {e}\n\nMake sure Chrome + chromedriver are installed.")

# ─── Tab 2: Bulk CSV ──────────────────────────────────────────────────────────
with tab2:
    st.subheader("Look up many products")
    st.markdown(
        "Upload a CSV with a column named **`ASIN`**. "
        "You'll get back an Excel file with titles + clickable links."
    )

    with st.expander("📋 Sample CSV format"):
        sample = pd.DataFrame({"ASIN": ["B0CHX3QBCH", "B09G9HD6PD", "B08N5WRWNW"]})
        st.dataframe(sample, use_container_width=True)
        sample_csv = sample.to_csv(index=False).encode()
        st.download_button("Download sample CSV", sample_csv, "sample_asins.csv", "text/csv")

    uploaded = st.file_uploader("Upload CSV", type=["csv"])

    if uploaded:
        df = pd.read_csv(uploaded, dtype=str)

        if "ASIN" not in df.columns:
            st.error("❌ CSV must have a column named **ASIN** (case-sensitive).")
        else:
            asins = df["ASIN"].dropna().str.strip().str.upper().tolist()
            total = len(asins)
            st.info(f"Found **{total} ASINs** in your file.")

            if st.button("🚀 Start Fetching", type="primary", key="bulk_btn"):
                results = []
                progress_bar = st.progress(0)
                status_text  = st.empty()
                found_count  = 0

                try:
                    driver = get_driver()

                    for i, asin in enumerate(asins, 1):
                        url   = f"https://www.amazon.in/dp/{asin}"
                        title = fetch_title(driver, asin)
                        results.append({"ASIN": asin, "Amazon URL": url, "Product Title": title})

                        if title != "Title not found":
                            found_count += 1

                        progress_bar.progress(i / total)
                        status_text.markdown(
                            f"Processing **{i}/{total}** — "
                            f"{'✅' if title != 'Title not found' else '❌'} `{asin}`"
                        )

                        # batch pause every 50
                        if i % 50 == 0 and i < total:
                            status_text.markdown(f"⏸ Pausing 15s to avoid blocks… ({i}/{total} done)")
                            time.sleep(15)

                    driver.quit()

                except Exception as e:
                    st.error(f"Chrome error: {e}")
                    st.stop()

                progress_bar.empty()
                status_text.empty()

                st.success(f"✅ Done! **{found_count}/{total}** titles fetched successfully.")

                excel_bytes = build_excel(results)
                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_bytes,
                    file_name="amazon_products_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # preview
                with st.expander("Preview results"):
                    st.dataframe(
                        pd.DataFrame(results)[["ASIN", "Product Title"]],
                        use_container_width=True,
                    )

# ── footer ────────────────────────────────────────────────────────────────────
st.divider()
st.caption("Tip: Amazon may occasionally block requests. If titles come back empty, wait 10–15 minutes and try again with a smaller batch.")
